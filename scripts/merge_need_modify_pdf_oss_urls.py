#!/usr/bin/env python3
"""将「素材ID → 样例 PDF 的 OSS URL」写入需修改相关表。

- **需修改分析**（单表 xlsx）：写 `样例PDF_OSS_URL`（默认仍更新此文件）。
- **家乐 workflow 总表与分表**：只改「需修改」sheet（用 openpyxl，避免整簿重写破坏「看板数据」等格式）。
"""
from __future__ import annotations

import argparse
import re
from pathlib import Path

import pandas as pd
from openpyxl import load_workbook

DEFAULT_ANALYSIS = Path("output/run_0412_0512/need_modify_analysis/需修改分析_Gemini双脚本.xlsx")
DEFAULT_WORKFLOW = Path("output/run_0412_0512/家乐_workflow_0412-0512_总表与分表.xlsx")
COL = "样例PDF_OSS_URL"


def normalize_id(x: object) -> str:
    s = str(x).strip()
    if s.endswith(".0") and s[:-2].isdigit():
        s = s[:-2]
    return s


def load_mapping_tsv(path: Path) -> dict[str, str]:
    out: dict[str, str] = {}
    for line in path.read_text(encoding="utf-8").splitlines():
        line = line.strip()
        if not line or line.startswith("#"):
            continue
        parts = line.split("\t")
        if len(parts) < 2:
            continue
        key = parts[0].strip()
        url = parts[1].strip()
        if key.endswith(".pdf"):
            key = key[: -len(".pdf")]
        if url.startswith("http"):
            out[normalize_id(key)] = url
    return out


def load_mapping_paste(path: Path) -> dict[str, str]:
    out: dict[str, str] = {}
    for line in path.read_text(encoding="utf-8").splitlines():
        line = line.strip()
        m = re.match(r"^\d+\t(.+\.pdf)\t(https://\S+)", line)
        if not m:
            continue
        fn, url = m.group(1), m.group(2)
        mid = fn.replace(".pdf", "").strip()
        out[normalize_id(mid)] = url
    return out


def build_auto_urls(base: str, material_ids: list[str]) -> dict[str, str]:
    b = base.rstrip("/") + "/"
    return {mid: f"{b}{mid}.pdf" for mid in material_ids}


def reorder_with_col(df: pd.DataFrame, col: str, anchor: str) -> pd.DataFrame:
    cols = list(df.columns)
    if col in cols:
        cols.remove(col)
    if anchor in cols:
        cols.insert(cols.index(anchor) + 1, col)
    else:
        cols.append(col)
    return df[cols]


def merge_urls_into_df(df: pd.DataFrame, mp: dict[str, str], col: str) -> pd.DataFrame:
    out = df.copy()
    out["素材ID"] = out["素材ID"].map(normalize_id)
    out[col] = out["素材ID"].map(lambda i: mp.get(normalize_id(i), ""))
    return out


def print_stats(df: pd.DataFrame, col: str, mp: dict[str, str]) -> None:
    missing = df[df[col].eq("")]["素材ID"].tolist()
    extra = sorted(set(mp.keys()) - set(df["素材ID"].tolist()))
    print(f"映射条数: {len(mp)}；表内匹配: {(df[col] != '').sum()}；表行数: {len(df)}")
    if missing:
        print(f"表中未匹配到 URL 的素材ID（{len(missing)}）: {missing[:20]}{'...' if len(missing) > 20 else ''}")
    if extra:
        print(f"映射里有但表中无（{len(extra)}）: {extra[:20]}{'...' if len(extra) > 20 else ''}")


def write_analysis_xlsx(path: Path, df: pd.DataFrame, col: str) -> None:
    anchor = "样例PDF本地路径" if "样例PDF本地路径" in df.columns else "gemini状态"
    df = reorder_with_col(df, col, anchor)
    path.parent.mkdir(parents=True, exist_ok=True)
    df.to_excel(path, index=False)
    csv_path = path.with_suffix(".csv")
    df.to_csv(csv_path, index=False, encoding="utf-8-sig")
    print(f"写入: {path}")
    print(f"写入: {csv_path}")


def update_workflow_need_modify_openpyxl(path: Path, mp: dict[str, str], col: str) -> None:
    """仅改「需修改」sheet，插入列在 oss_path 右侧；已有同名列则覆盖。"""
    path = path.resolve()
    wb = load_workbook(path)
    if "需修改" not in wb.sheetnames:
        raise SystemExit(f"{path.name} 中无「需修改」sheet")
    ws = wb["需修改"]
    header_row = 1
    header_map: dict[str, int] = {}
    for c in range(1, ws.max_column + 1):
        v = ws.cell(header_row, c).value
        if v is not None and str(v).strip():
            header_map[str(v).strip()] = c
    if "素材ID" not in header_map:
        raise SystemExit("需修改表无「素材ID」列")
    mid_col = header_map["素材ID"]
    if col in header_map:
        url_col = header_map[col]
    elif "oss_path" in header_map:
        url_col = header_map["oss_path"] + 1
        ws.insert_cols(url_col)
        ws.cell(header_row, url_col, col)
    else:
        url_col = ws.max_column + 1
        ws.cell(header_row, url_col, col)

    for r in range(2, ws.max_row + 1):
        mid = normalize_id(ws.cell(r, mid_col).value)
        ws.cell(r, url_col, mp.get(mid, ""))

    wb.save(path)
    print(f"写入 workflow「需修改」sheet（openpyxl，其它 sheet 未动）: {path}")


def main() -> None:
    ap = argparse.ArgumentParser(description=__doc__)
    ap.add_argument(
        "--xlsx",
        "--analysis-xlsx",
        dest="analysis_xlsx",
        type=Path,
        default=None,
        help=f"需修改分析单表（默认: {DEFAULT_ANALYSIS}；与 --skip-analysis 互斥）",
    )
    ap.add_argument(
        "--workflow-xlsx",
        type=Path,
        default=None,
        help=f"家乐 workflow 总表与分表（常用: {DEFAULT_WORKFLOW}）",
    )
    ap.add_argument(
        "--skip-analysis",
        action="store_true",
        help="不更新需修改分析 xlsx/csv",
    )
    ap.add_argument("--map", type=Path, help="TSV：素材ID 或 素材ID.pdf + TAB + URL")
    ap.add_argument("--paste", type=Path, help="「序号\\txxx.pdf\\turl」粘贴文本文件")
    ap.add_argument(
        "--auto-oss-base",
        metavar="URL_PREFIX",
        help="OSS 目录 URL 前缀，每条生成 PREFIX+素材ID+.pdf",
    )
    args = ap.parse_args()

    if not args.auto_oss_base and not args.paste and not args.map:
        raise SystemExit("请指定 --auto-oss-base、--paste 或 --map")

    workflow_path = args.workflow_xlsx.resolve() if args.workflow_xlsx else None

    analysis_path: Path | None = None
    if not args.skip_analysis:
        analysis_path = (args.analysis_xlsx or DEFAULT_ANALYSIS).resolve()
        if not analysis_path.exists():
            print(f"[warn] 跳过分析表（不存在）: {analysis_path}", flush=True)
            analysis_path = None

    id_lists: list[list[str]] = []
    if workflow_path and workflow_path.exists():
        mod = pd.read_excel(workflow_path, sheet_name="需修改", dtype=str)
        id_lists.append([normalize_id(i) for i in mod["素材ID"].tolist()])
    if analysis_path and analysis_path.exists():
        adf = pd.read_excel(analysis_path, dtype=str)
        id_lists.append([normalize_id(i) for i in adf["素材ID"].tolist()])

    if not id_lists:
        raise SystemExit("未找到 workflow 或 analysis：请检查 --workflow-xlsx / 默认分析路径")

    all_ids = sorted({i for lst in id_lists for i in lst})

    if args.auto_oss_base:
        mp = build_auto_urls(args.auto_oss_base, all_ids)
    elif args.paste:
        mp = load_mapping_paste(args.paste.resolve())
    else:
        mp = load_mapping_tsv(args.map.resolve())

    did = False
    if workflow_path:
        if not workflow_path.exists():
            raise SystemExit(f"找不到 workflow: {workflow_path}")
        mod = pd.read_excel(workflow_path, sheet_name="需修改", dtype=str)
        mod_m = merge_urls_into_df(mod, mp, COL)
        print_stats(mod_m, COL, mp)
        update_workflow_need_modify_openpyxl(workflow_path, mp, COL)
        did = True

    if analysis_path and analysis_path.exists():
        df0 = pd.read_excel(analysis_path, dtype=str)
        df0["素材ID"] = df0["素材ID"].map(normalize_id)
        df = merge_urls_into_df(df0, mp, COL)
        print_stats(df, COL, mp)
        write_analysis_xlsx(analysis_path, df, COL)
        did = True

    if not did:
        raise SystemExit("未写入任何文件：请指定存在的 --workflow-xlsx，或勿使用 --skip-analysis 并保证分析表存在")


if __name__ == "__main__":
    main()
