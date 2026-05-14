#!/usr/bin/env python3
"""看板数据 sheet（分段+合并商品列+结论区）与编导向 MD 分析报告。"""
from __future__ import annotations

from dataclasses import dataclass
from pathlib import Path
from typing import Any, Dict, List, Sequence, Tuple

import pandas as pd
from openpyxl import Workbook
from openpyxl.styles import Alignment, Border, Font, PatternFill, Side
from openpyxl.utils import get_column_letter

# 看板行顺序、MD 主章节顺序、结论区按品汇总顺序（与编导习惯一致；小面酱与三品同级）
MAIN_REPORT_PRODUCTS = ["回味粉", "干锅酱", "水煮酱", "小面酱"]


def _pct_share(n: int, total: int) -> str:
    if total <= 0:
        return "0（0%）"
    p = 100.0 * n / total
    return f"{n}（{p:.0f}%）"


def add_three_class(detail: pd.DataFrame, wf: Any) -> pd.DataFrame:
    out = detail.copy()
    out["三类结果"] = out.apply(lambda r: wf.classify_result(str(r["商品"]), r["GMV"]), axis=1)
    return out


def _weighted(sub: pd.DataFrame, rate_col: str, wcol: str = "播放") -> float | None:
    t = sub[[rate_col, wcol]].dropna()
    t = t[(t[wcol] > 0) & t[rate_col].notna()]
    if t.empty:
        return None
    return float((t[rate_col] * t[wcol]).sum() / t[wcol].sum())


def build_board_rows(detail: pd.DataFrame, wf: Any) -> pd.DataFrame:
    """按商品×账号（奥创、轩辕、总计）+ 全部总计，含三分类数量占比。"""
    rows: List[Dict[str, Any]] = []
    present = set(detail["商品"].dropna().unique())
    target = set(getattr(wf, "TARGET_PRODUCTS", MAIN_REPORT_PRODUCTS))
    products = [p for p in MAIN_REPORT_PRODUCTS if p in present and p in target]
    for p in sorted(present & target):
        if p not in products:
            products.append(p)
    if not products and not detail.empty:
        products = sorted(detail["商品"].dropna().unique().tolist())

    def one_block(sub_all: pd.DataFrame) -> None:
        for product in products:
            p_df = sub_all[sub_all["商品"] == product]
            if p_df.empty:
                continue
            for account in ["奥创", "轩辕"]:
                sub = p_df[p_df["账号"] == account]
                if sub.empty:
                    continue
                rows.append(_row_metrics(sub, product, account, wf))
            rows.append(_row_metrics(p_df, product, "总计", wf))
        if not sub_all.empty:
            rows.append(_row_metrics(sub_all, "全部", "总计", wf))

    one_block(detail)
    return pd.DataFrame(rows)


def _row_metrics(sub: pd.DataFrame, product: str, account: str, wf: Any) -> Dict[str, Any]:
    vc = int(sub["素材ID"].nunique())
    gmv = float(sub["GMV"].fillna(0).sum())
    spend = float(sub["消耗"].fillna(0).sum())
    play = float(sub["播放"].fillna(0).sum())
    show = float(sub["展示"].fillna(0).sum())
    click = float(sub["点击"].fillna(0).sum())
    order = float(sub["订单"].fillna(0).sum())
    cls = sub["三类结果"].value_counts()
    n_copy = int(cls.get("可复制", 0))
    n_mod = int(cls.get("需修改", 0))
    n_drop = int(cls.get("直接放弃", 0))
    tot = n_copy + n_mod + n_drop
    if tot != vc and vc > 0:
        tot = vc
    roi = (gmv / spend) if spend else None
    ctr = (click / show) if show else None
    cvr = (order / click) if click else None
    return {
        "商品": product,
        "账号": account,
        "视频条数": vc,
        "总GMV": gmv,
        "平均GMV": (gmv / vc) if vc else None,
        "总消耗": spend,
        "平均消耗": (spend / vc) if vc else None,
        "总播放": play,
        "平均播放": (play / vc) if vc else None,
        "总ROI": roi,
        "加权CTR": ctr,
        "加权CVR": cvr,
        "加权3s完播率": _weighted(sub, "3s完播率", "播放"),
        "加权5s完播率": _weighted(sub, "5s完播率", "播放"),
        "加权10s完播率": _weighted(sub, "10s完播率", "播放"),
        "可复制": _pct_share(n_copy, vc),
        "需修改": _pct_share(n_mod, vc),
        "直接放弃": _pct_share(n_drop, vc),
    }


def monday_week(ts: pd.Timestamp) -> pd.Timestamp:
    d = ts.normalize()
    return d - pd.Timedelta(days=int(d.weekday()))


def week_spans_in_range(start: pd.Timestamp, end: pd.Timestamp) -> List[Tuple[pd.Timestamp, pd.Timestamp, str]]:
    """自然周 周一至周日，列出与 [start,end] 有交集的周。"""
    out: List[Tuple[pd.Timestamp, pd.Timestamp, str]] = []
    s0 = start.normalize()
    e0 = end.normalize()
    cur = monday_week(s0)
    while cur <= e0:
        wend = cur + pd.Timedelta(days=6)
        if wend < s0:
            cur += pd.Timedelta(days=7)
            continue
        label = f"{cur.strftime('%m%d')}-{wend.strftime('%m%d')}"
        out.append((cur, wend, label))
        cur += pd.Timedelta(days=7)
    return out


def format_board_excel_values(df: pd.DataFrame) -> pd.DataFrame:
    """CTR/CVR/完播为两位小数比例（写 Excel 时用百分比格式）；ROI 两位小数；其余整数。"""
    out = df.copy()
    int_cols = ["视频条数", "总GMV", "平均GMV", "总消耗", "平均消耗", "总播放", "平均播放"]
    for c in int_cols:
        if c in out.columns:
            out[c] = pd.to_numeric(out[c], errors="coerce").round(0)
    if "总ROI" in out.columns:
        out["总ROI"] = pd.to_numeric(out["总ROI"], errors="coerce").round(2)
    for c in ["加权CTR", "加权CVR", "加权3s完播率", "加权5s完播率", "加权10s完播率"]:
        if c in out.columns:
            out[c] = pd.to_numeric(out[c], errors="coerce").round(4)
    return out


@dataclass
class DashboardSection:
    title: str
    df: pd.DataFrame


def _thin() -> Side:
    return Side(style="thin", color="CCCCCC")


def write_dashboard_sheet(
    wb: Workbook,
    sheet_name: str,
    sections: List[DashboardSection],
    conclusion: str,
    data_start_col_conclusion: int = 18,
) -> None:
    ws = wb.create_sheet(sheet_name)
    wrap = Alignment(wrap_text=True, vertical="top")
    bold = Font(bold=True)
    fill = PatternFill("solid", fgColor="D9EAF7")
    title_font = Font(bold=True, size=11)
    row_idx = 1
    headers = None
    for sec in sections:
        ws.merge_cells(start_row=row_idx, start_column=1, end_row=row_idx, end_column=16)
        c = ws.cell(row_idx, 1, sec.title)
        c.font = title_font
        c.alignment = wrap
        row_idx += 1
        df = sec.df.copy()
        if df.empty:
            ws.cell(row_idx, 1, "（无数据）")
            row_idx += 2
            continue
        headers = list(df.columns)
        for j, h in enumerate(headers, 1):
            cell = ws.cell(row_idx, j, h)
            cell.font = bold
            cell.fill = fill
            cell.alignment = wrap
            cell.border = Border(bottom=_thin())
        header_row = row_idx
        row_idx += 1
        data_start = row_idx
        for _, r in df.iterrows():
            for j, h in enumerate(headers, 1):
                val = r[h]
                cell = ws.cell(row_idx, j, val)
                cell.alignment = wrap
                if h in ("加权CTR", "加权CVR", "加权3s完播率", "加权5s完播率", "加权10s完播率") and isinstance(val, (int, float)) and pd.notna(val):
                    cell.number_format = "0.00%"
                elif h == "总ROI" and isinstance(val, (int, float)) and pd.notna(val):
                    cell.number_format = "0.00"
                elif h in ("视频条数", "总GMV", "平均GMV", "总消耗", "平均消耗", "总播放", "平均播放") and isinstance(val, (int, float)) and pd.notna(val):
                    cell.number_format = "#,##0"
                cell.border = Border(bottom=_thin(), right=_thin())
            row_idx += 1
        data_end = row_idx - 1
        # 合并「商品」列连续相同块（奥创/轩辕/总计）
        if headers and "商品" in headers:
            col_idx = headers.index("商品") + 1
            i = data_start
            while i <= data_end:
                val = ws.cell(i, col_idx).value
                j = i + 1
                while j <= data_end and ws.cell(j, col_idx).value == val:
                    j += 1
                if j - 1 > i:
                    ws.merge_cells(start_row=i, start_column=col_idx, end_row=j - 1, end_column=col_idx)
                    ws.cell(i, col_idx).alignment = Alignment(vertical="center", wrap_text=True)
                i = j
        row_idx += 1

    # 结论区
    c0 = data_start_col_conclusion
    ws.merge_cells(start_row=1, start_column=c0, end_row=max(80, row_idx + 25), end_column=c0 + 5)
    cc = ws.cell(1, c0, conclusion.strip())
    cc.alignment = Alignment(wrap_text=True, vertical="top")
    cc.font = Font(size=10)

    for col in range(1, 18):
        ws.column_dimensions[get_column_letter(col)].width = 14
    ws.column_dimensions[get_column_letter(c0)].width = 22


def build_conclusion_text(
    detail_month_create: pd.DataFrame,
    week_summaries: List[Tuple[str, pd.DataFrame]],
    spend_range: str,
    create_range: str,
    wf: Any,
) -> str:
    lines: List[str] = []
    lines.append("【数据结论】（供快速对照；细拆见下方 MD 报告）")
    lines.append("")
    lines.append(f"- 月消耗看板统计口径：消耗时间 {spend_range}；创建时间：全部创建时间（当前导出内可见的全部素材）。")
    lines.append(f"- 月创建看板统计口径：消耗时间 {spend_range}；创建时间 {create_range}（与看板/总体明细约定的创建窗口一致）。")
    lines.append("- 周看板：按自然周（周一至周日）切分创建时间；最新一周若未满 7 天，跑量与转化可能仍会上调，结论偏保守。")
    lines.append("")
    if detail_month_create.empty:
        lines.append("月创建样本为空，无法对比奥创/轩辕。")
        return "\n".join(lines)

    def acc_summary(d: pd.DataFrame, label: str) -> None:
        lines.append(f"### {label}")
        for p in MAIN_REPORT_PRODUCTS:
            if p not in getattr(wf, "TARGET_PRODUCTS", MAIN_REPORT_PRODUCTS):
                continue
            sub = d[d["商品"] == p]
            if sub.empty:
                continue
            for acc in ["奥创", "轩辕"]:
                a = sub[sub["账号"] == acc]
                if a.empty:
                    continue
                vc = a["素材ID"].nunique()
                g = a["GMV"].fillna(0).sum()
                avg = g / vc if vc else 0
                cls = a["三类结果"].value_counts()
                lines.append(
                    f"- {p}｜{acc}：素材 {int(vc)} 条，总 GMV {g:,.0f}，平均 GMV {avg:,.1f}；"
                    f"可复制 {int(cls.get('可复制', 0))}、需修改 {int(cls.get('需修改', 0))}、直接放弃 {int(cls.get('直接放弃', 0))}"
                )
        lines.append("")

    acc_summary(detail_month_create, "当月创建（主结论口径）各品各账号条数与 GMV、三分类数量")

    lines.append("### 关键指标谁更强（当月创建、按账号汇总后粗比）")
    agg = []
    for acc in ["奥创", "轩辕"]:
        a = detail_month_create[detail_month_create["账号"] == acc]
        if a.empty:
            continue
        show = a["展示"].fillna(0).sum()
        click = a["点击"].fillna(0).sum()
        order = a["订单"].fillna(0).sum()
        play = a["播放"].fillna(0).sum()
        ctr = click / show if show else None
        cvr = order / click if click else None
        w3 = _weighted(a, "3s完播率", "播放")
        agg.append((acc, ctr, cvr, w3, a["GMV"].fillna(0).sum()))
    if len(agg) == 2:
        (a1, c1, v1, s1, g1), (a2, c2, v2, s2, g2) = agg[0], agg[1]

        def cmp_metric(name: str, x1, x2, higher_better: bool) -> str:
            if x1 is None or x2 is None:
                return f"{name}：数据不足"
            better = a1 if (x1 > x2) == higher_better else a2
            worse = a2 if better == a1 else a1
            return f"{name}：{better} 更优（约 {max(x1, x2) / max(min(x1, x2), 1e-9):.2f} 倍相对另一账号）"

        lines.append(
            cmp_metric("CTR（偏点击端）", c1 or 0, c2 or 0, True)
            if (c1 or c2)
            else "CTR：数据不足"
        )
        lines.append(
            cmp_metric("CVR（偏转化端）", v1 or 0, v2 or 0, True)
            if (v1 or v2)
            else "CVR：数据不足"
        )
        lines.append(
            cmp_metric("加权 3s 完播（偏开头承接）", s1 or 0, s2 or 0, True)
            if (s1 or s2)
            else "3s 完播：数据不足"
        )
        lines.append(f"- 总 GMV：{a1 if g1 >= g2 else a2} 更高（{max(g1, g2):,.0f} vs {min(g1, g2):,.0f}）")
    lines.append("")
    lines.append("### 周度提示")
    for lab, wdf in week_summaries:
        if wdf.empty:
            lines.append(f"- 【{lab}】：无素材")
            continue
        lines.append(f"- 【{lab}】：素材 {wdf['素材ID'].nunique()} 条，总 GMV {wdf['GMV'].fillna(0).sum():,.0f}")
    lines.append("")
    lines.append("可复制阈值：回味粉 GMV≥500；干锅酱/水煮酱/小面酱 GMV≥100。若某账号「可复制」占比高且总 GMV也高，需警惕爆款拉高平均 GMV。")
    return "\n".join(lines)


def _top_copy_in_month(d: pd.DataFrame, product: str, account: str, n: int = 2) -> pd.DataFrame:
    sub = d[(d["商品"] == product) & (d["账号"] == account) & (d["三类结果"] == "可复制")]
    sub = sub.sort_values("GMV", ascending=False)
    return sub.head(n)


def build_md_report(
    detail_month_create: pd.DataFrame,
    week_summaries: List[Tuple[str, pd.DataFrame]],
    spend_range: str,
    create_range: str,
    wf: Any,
    need_modify_xlsx: Path | None,
) -> str:
    """编导向：按回味粉、干锅酱、水煮酱、小面酱四节。"""
    lines: List[str] = []
    lines.append("# 家乐千川编导读数简报（以「月创建看板」为主口径）")
    lines.append("")
    lines.append(f"- **消耗时间范围**：{spend_range}")
    lines.append(f"- **创建时间范围（月创建）**：{create_range}")
    lines.append("- **周看板**：按周一至周日；**最后一周若未满 7 天**，数据可能仍变化，结论仅供参考。")
    lines.append("")
    lines.append("---")
    lines.append("")

    nm: Dict[str, str] = {}
    if need_modify_xlsx and need_modify_xlsx.exists():
        try:
            ndf = pd.read_excel(need_modify_xlsx, dtype=str)
            for _, r in ndf.iterrows():
                mid = wf.normalize_id(r.get("素材ID"))
                if mid:
                    nm[mid] = wf.safe_str(r.get("具体修改建议", ""))[:400]
        except Exception:
            pass

    def section_product(product: str) -> None:
        lines.append(f"## {product}")
        d = detail_month_create[detail_month_create["商品"] == product]
        if d.empty:
            lines.append("本月创建窗口内无该品素材。")
            lines.append("")
            return
        lines.append("### 编导一眼结论（少数字，多看方向）")
        if product == "回味粉":
            lines.append(
                "- 回味粉：爆款里高频出现「汤底鲜香 / 回头客 / 对比鸡精味精」三件事，轩辕若可复制更多，优先抄它的开头「痛点一句话 + 立刻给方案」节奏。"
            )
        elif product == "干锅酱":
            lines.append("- 干锅酱：锅气画面 +「一锅出菜、省料省时」比空讲辣更吃香；夜市/门店场景更容易让老板代入。")
        elif product == "水煮酱":
            lines.append("- 水煮酱：抓住「肉片嫩滑 + 红油食欲 + 出餐速度」，避免只讲辣不讲香。")
        elif product == "小面酱":
            lines.append(
                "- 小面酱：抓住「重庆小面/面馆场景 + 一包底料定味 + 出餐快」；画面给红油、挑面、嗦面，比空讲配方更带货。"
            )
        else:
            lines.append("- 结合可复制素材的标题与 ASR，找「主菜明确 + 卖点一句顶穿」的共性。")
        lines.append("")
        lines.append("### 整体谁多、谁更赚钱（当月创建）")
        for acc in ["轩辕", "奥创"]:
            a = d[d["账号"] == acc]
            if a.empty:
                lines.append(f"- **{acc}**：无素材")
                continue
            vc = int(a["素材ID"].nunique())
            g = float(a["GMV"].fillna(0).sum())
            avg = g / vc if vc else 0.0
            cls = a["三类结果"].value_counts()
            nc, nm_, nd = int(cls.get("可复制", 0)), int(cls.get("需修改", 0)), int(cls.get("直接放弃", 0))
            lines.append(
                f"- **{acc}**：{vc} 条；总 GMV **{g:,.0f}**；平均 GMV **{avg:,.1f}**；"
                f"可复制 **{nc}**、需修改 **{nm_}**、直接放弃 **{nd}**。"
            )
        lines.append("")
        lines.append("### 周度：哪一周更好（粗看条数+GMV）")
        for lab, wdf in week_summaries:
            wp = wdf[wdf["商品"] == product]
            if wp.empty:
                lines.append(f"- **{lab}**：无该品素材")
                continue
            g = wp["GMV"].fillna(0).sum()
            lines.append(
                f"- **{lab}**：{int(wp['素材ID'].nunique())} 条，总 GMV {g:,.0f}（轩辕 {int(wp[wp['账号']=='轩辕']['素材ID'].nunique())} / 奥创 {int(wp[wp['账号']=='奥创']['素材ID'].nunique())}）"
            )
        lines.append("")
        lines.append("### 内容方向（结合三分类与可复制 ASR）")
        lines.append(
            f"- 若轩辕「可复制」明显多于奥创，而奥创「直接放弃」更多：优先对比**开头钩子、菜品选题是否更贴 B 端痛点**、以及是否过度依赖单一菜式。"
        )
        lines.append("")
        lines.append("### 当月「可复制」代表爆款（含完整 ASR，便于拆框架）")
        for acc in ["轩辕", "奥创"]:
            tops = _top_copy_in_month(detail_month_create, product, acc, 2)
            if tops.empty:
                lines.append(f"#### {acc}：当月无可复制爆款")
                lines.append("")
                continue
            for _, r in tops.iterrows():
                mid = wf.normalize_id(r["素材ID"])
                lines.append(f"#### {acc}｜GMV {float(r.get('GMV', 0) or 0):,.1f}｜{wf.safe_str(r.get('素材名称'))[:60]}")
                lines.append(f"- **素材ID**：`{mid}`")
                lines.append(f"- **oss_path**：{wf.safe_str(r.get('oss_path'))}")
                asr = wf.safe_str(r.get("asr", ""))
                lines.append("- **ASR（完整）**")
                lines.append("")
                lines.append(asr if asr else "（无 ASR）")
                lines.append("")
                if mid in nm and nm[mid]:
                    lines.append("- **需修改侧参考（若有同 ID 分析）**：")
                    lines.append(nm[mid])
                    lines.append("")
        lines.append("### 可复制框架共同点（编导执行清单）")
        lines.append("- **开头**：前 3 秒点名「老板/后厨痛点」+ 明确品类场景，少铺垫。")
        if product == "小面酱":
            lines.append(
                "- **中段**：用「一碗定味 / 与自家熬料对比」讲清省时省力；强调麻辣鲜香层次、汤底挂面、复购话术。"
            )
        else:
            lines.append("- **中段**：用「对比鸡精味精 / 单一调味」制造差异，把鲜香、回味、出餐效率讲具体。")
        lines.append("- **结尾**：给清晰动作（左下角/新客福利/试用门槛），降低决策成本。")
        if product == "回味粉":
            lines.append("- **回味粉加练**：用「三天不喝就惦记」「一锅汤顶一晚熬」这类**可复述金句**；把「回头粉」与「回头率」做谐音记忆。")
        elif product == "干锅酱":
            lines.append("- **干锅酱加练**：镜头多给「翻锅/浇油/冒泡」；话术强调「一锅成菜、少备料、出品稳定」。")
        elif product == "水煮酱":
            lines.append("- **水煮酱加练**：先讲「嫩肉不碎」再给「红油香而不呛」；结尾用「一锅出餐多少份」帮老板算账。")
        elif product == "小面酱":
            lines.append(
                "- **小面酱加练**：特写「红油亮、花椒麻、豌豆酥」；用「早高峰出 200 碗也不乱味」类**可量化**话术增强信任。"
            )
        lines.append("")
        lines.append("---")
        lines.append("")

    for p in MAIN_REPORT_PRODUCTS:
        section_product(p)

    return "\n".join(lines)


def prepare_detail_numeric(raw: pd.DataFrame, oss: pd.DataFrame, wf: Any) -> pd.DataFrame:
    """与流水线一致：合并 OSS/ASR，保留数值列。"""
    m = raw.merge(oss, on="素材ID", how="left")
    m["asr"] = m["asr_prefill"].fillna("").map(lambda x: wf.clean_text(x))
    m = m.drop(columns=["asr_prefill"], errors="ignore")
    m["oss_path"] = m["oss_path"].fillna("").astype(str)
    return m


def format_master_detail_no_asr(detail: pd.DataFrame, wf: Any) -> pd.DataFrame:
    """总体明细表：不写 asr，时间格式字符串。"""
    out = detail.copy()
    if "素材创建时间" in out.columns and pd.api.types.is_datetime64_any_dtype(out["素材创建时间"]):
        out["素材创建时间"] = out["素材创建时间"].dt.strftime("%Y-%m-%d %H:%M:%S")
    out["三类结果"] = out.apply(lambda r: wf.classify_result(str(r["商品"]), r["GMV"]), axis=1)
    cols = [
        "商品",
        "账号",
        "素材ID",
        "素材名称",
        "素材创建时间",
        "发布时间",
        "oss_path",
        "消耗",
        "GMV",
        "ROI",
        "播放",
        "千次播放GMV",
        "展示",
        "点击",
        "订单",
        "CTR",
        "CVR",
        "3s完播率",
        "5s完播率",
        "10s完播率",
        "三类结果",
    ]
    for c in cols:
        if c not in out.columns:
            out[c] = pd.NA
    return out[cols].copy()


def _style_header_row(ws: Any, row: int, ncols: int) -> None:
    wrap = Alignment(wrap_text=True, vertical="top")
    bold = Font(bold=True)
    fill = PatternFill("solid", fgColor="D9EAF7")
    for j in range(1, ncols + 1):
        cell = ws.cell(row, j)
        cell.font = bold
        cell.fill = fill
        cell.alignment = wrap


def write_dataframe_sheet(wb: Workbook, name: str, df: pd.DataFrame) -> None:
    ws = wb.create_sheet(name)
    if df.empty:
        ws.cell(1, 1, "（无数据）")
        return
    cols = list(df.columns)
    for j, h in enumerate(cols, 1):
        ws.cell(1, j, h)
    _style_header_row(ws, 1, len(cols))
    wrap = Alignment(wrap_text=True, vertical="top")
    for i, (_, r) in enumerate(df.iterrows(), start=2):
        for j, h in enumerate(cols, 1):
            c = ws.cell(i, j, r[h])
            c.alignment = wrap
    for j in range(1, len(cols) + 1):
        ws.column_dimensions[get_column_letter(j)].width = min(48, 14)


def write_step1_workbook_v2(
    path: Path,
    detail_master_fmt: pd.DataFrame,
    fmt_win: pd.DataFrame,
    fmt_hist: pd.DataFrame,
    copy_df: pd.DataFrame,
    modify_df: pd.DataFrame,
    drop_df: pd.DataFrame,
    sections: List[DashboardSection],
    conclusion: str,
) -> None:
    wb = Workbook()
    wb.remove(wb.active)
    write_dataframe_sheet(wb, "总体明细表", detail_master_fmt)
    write_dataframe_sheet(wb, "总明细表", fmt_win)
    write_dataframe_sheet(wb, "历史跑量爆款", fmt_hist)
    write_dataframe_sheet(wb, "可复制", copy_df)
    write_dataframe_sheet(wb, "需修改", modify_df)
    write_dataframe_sheet(wb, "直接放弃", drop_df)
    write_dashboard_sheet(wb, "看板数据", sections, conclusion)
    wb.save(path)
