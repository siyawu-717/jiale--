#!/usr/bin/env python3
from __future__ import annotations

import argparse
import csv
import hashlib
import importlib.util
import json
import re
import time
import uuid
from concurrent.futures import ThreadPoolExecutor, as_completed
from pathlib import Path
from typing import Any, Dict, Iterable, List

import pandas as pd
from openpyxl import load_workbook
from openpyxl.styles import Alignment, Font, PatternFill


ROOT = Path("/Users/siya/Desktop/家乐/家乐千川数据分析")
AOCHUANG_XLSX = ROOT / "奥创最近30天.xlsx"
YUANYUAN_XLSX = ROOT / "轩辕最近30天.xlsx"
OSS_XLSX = ROOT / "上上周创建新素材oss.xlsx"
MAINTAIN_XLSX = ROOT / "信息维护表-test.xlsx"
OUTPUT_DIR = ROOT / "output" / "上上周新创建素材分析"
SKILL_CALL_GEMINI = Path("/Users/siya/.codex/skills/gemini-las-video/scripts/call_gemini.py")
PROXY_BASE = "https://las-operator.runix.ai"
TOS_BASE = "https://mogic-collect.tos-cn-beijing.volces.com"
MODEL_NAME = "gemini-2.5-flash"
TARGET_PRODUCTS = ["回味粉", "小面酱", "干锅酱", "水煮酱"]


def safe_str(value: Any) -> str:
    if value is None:
        return ""
    if isinstance(value, str):
        return value.strip()
    return str(value).strip()


def clean_text(value: Any) -> str:
    return re.sub(r"\s+", " ", safe_str(value)).strip()


def normalize_id(value: Any) -> str:
    text = safe_str(value)
    return text[:-2] if text.endswith(".0") else text


def parse_number(value: Any) -> float | None:
    text = safe_str(value).replace(",", "")
    if not text or text.lower() == "nan":
        return None
    try:
        return float(text)
    except ValueError:
        return None


def parse_percent(value: Any) -> float | None:
    text = safe_str(value)
    if not text or text.lower() == "nan":
        return None
    if text.endswith("%"):
        num = parse_number(text[:-1])
        return None if num is None else num / 100
    num = parse_number(text)
    if num is None:
        return None
    return num / 100 if 1 < num <= 100 else num


def infer_product(title: str) -> str | None:
    text = safe_str(title)
    if "回味粉" in text or "回头粉" in text:
        return "回味粉"
    # 小面酱：放在干锅/水煮之前，避免「水煮小面」等被误判为水煮酱
    if "小面酱" in text or "重庆小面" in text:
        return "小面酱"
    if "小面" in text and "小面筋" not in text:
        return "小面酱"
    if "干锅酱" in text or "干锅" in text:
        return "干锅酱"
    if "水煮酱" in text or "水煮麻辣" in text or "水煮" in text:
        return "水煮酱"
    return None


def unique_nonempty(items: Iterable[str]) -> List[str]:
    seen = set()
    result: List[str] = []
    for item in items:
        text = clean_text(item)
        if not text or text in seen:
            continue
        seen.add(text)
        result.append(text)
    return result


def load_recent30(path: Path, account: str) -> pd.DataFrame:
    df = pd.read_excel(path, dtype=str)
    name_col = "素材视频名称" if "素材视频名称" in df.columns else "素材名称"
    df = df.rename(columns={name_col: "素材名称"}).copy()
    df["账号"] = account
    df["素材ID"] = df["素材ID"].map(normalize_id)
    df["素材名称"] = df["素材名称"].map(clean_text)
    df["日期_raw"] = df["日期"].map(safe_str)
    df["日期"] = pd.to_datetime(df["日期_raw"], errors="coerce")
    df["素材创建时间"] = pd.to_datetime(df["素材创建时间"], errors="coerce")
    df["商品"] = df["素材名称"].map(infer_product)
    keep_cols = [
        "素材ID",
        "素材名称",
        "素材创建时间",
        "日期_raw",
        "日期",
        "账号",
        "商品",
        "整体展示次数",
        "整体点击次数",
        "整体点击率",
        "整体转化率",
        "整体消耗",
        "整体成交订单数",
        "整体成交金额",
        "视频播放数",
        "3秒播放率",
        "5秒播放率",
        "10秒播放率",
    ]
    for col in keep_cols:
        if col not in df.columns:
            df[col] = pd.NA
    return df[keep_cols].copy()


def load_maintenance() -> pd.DataFrame:
    if not MAINTAIN_XLSX.exists():
        return pd.DataFrame(columns=["素材ID", "账号", "奥创视频制作名称", "oss_path", "商品"])
    df = pd.read_excel(MAINTAIN_XLSX, dtype=str)
    df["素材ID"] = df["素材ID"].map(normalize_id)
    rename_map = {}
    if "奥创制作视频名称" in df.columns:
        rename_map["奥创制作视频名称"] = "奥创视频制作名称"
    df = df.rename(columns=rename_map).copy()
    for col in ["账号", "奥创视频制作名称", "oss_path", "商品"]:
        if col not in df.columns:
            df[col] = pd.NA
    df = df.sort_values(["素材ID"]).drop_duplicates(subset=["素材ID"], keep="first")
    return df[["素材ID", "账号", "奥创视频制作名称", "oss_path", "商品"]].copy()


def load_oss_map() -> tuple[pd.DataFrame, int]:
    df = pd.read_excel(OSS_XLSX, dtype=str)
    df["视频ID"] = df["视频ID"].map(normalize_id)
    df["OSS路径"] = df["OSS路径"].map(clean_text)
    duplicate_count = int(df.duplicated(subset=["视频ID"], keep="first").sum())
    df = df.drop_duplicates(subset=["视频ID"], keep="first")
    return df.rename(columns={"视频ID": "素材ID", "OSS路径": "oss_path"})[["素材ID", "oss_path"]].copy(), duplicate_count


def apply_maintenance(raw: pd.DataFrame, maintain: pd.DataFrame) -> pd.DataFrame:
    merged = raw.merge(
        maintain.rename(
            columns={
                "账号": "维护账号",
                "奥创视频制作名称": "维护奥创视频制作名称",
                "oss_path": "维护oss_path",
                "商品": "维护商品",
            }
        ),
        on="素材ID",
        how="left",
    )
    merged["账号"] = merged["维护账号"].where(merged["维护账号"].notna() & (merged["维护账号"].astype(str) != ""), merged["账号"])
    merged["商品"] = merged["商品"].where(merged["商品"].notna(), merged["维护商品"])
    merged["奥创视频制作名称"] = merged["维护奥创视频制作名称"]
    merged["维护oss_path"] = merged["维护oss_path"]
    return merged.drop(columns=["维护账号", "维护奥创视频制作名称", "维护商品"])


def weighted_rate(group: pd.DataFrame, rate_col: str, weight_col: str = "视频播放数_num") -> float | None:
    temp = group[[rate_col, weight_col]].dropna().copy()
    if temp.empty:
        return None
    den = temp[weight_col].sum()
    if not den:
        return None
    return float((temp[rate_col] * temp[weight_col]).sum() / den)


def aggregate_videos(raw: pd.DataFrame, publish_start: pd.Timestamp, publish_end: pd.Timestamp) -> pd.DataFrame:
    daily = raw[(raw["日期_raw"] != "全部") & raw["日期"].notna()].copy()
    daily = daily[daily["商品"].isin(TARGET_PRODUCTS)].copy()
    daily = daily[daily["素材创建时间"].dt.normalize().between(publish_start, publish_end, inclusive="both")].copy()
    if daily.empty:
        return daily

    numeric_map = {
        "整体展示次数_num": ("整体展示次数", parse_number),
        "整体点击次数_num": ("整体点击次数", parse_number),
        "整体消耗_num": ("整体消耗", parse_number),
        "整体成交订单数_num": ("整体成交订单数", parse_number),
        "整体成交金额_num": ("整体成交金额", parse_number),
        "视频播放数_num": ("视频播放数", parse_number),
        "3秒播放率_num": ("3秒播放率", parse_percent),
        "5秒播放率_num": ("5秒播放率", parse_percent),
        "10秒播放率_num": ("10秒播放率", parse_percent),
    }
    for new_col, (old_col, parser) in numeric_map.items():
        daily[new_col] = daily[old_col].map(parser)

    rows: List[Dict[str, Any]] = []
    key_cols = ["商品", "账号", "素材ID"]
    for (product, account, material_id), group in daily.groupby(key_cols, sort=False):
        spend = group["整体消耗_num"].fillna(0).sum()
        gmv = group["整体成交金额_num"].fillna(0).sum()
        plays = group["视频播放数_num"].fillna(0).sum()
        show = group["整体展示次数_num"].fillna(0).sum()
        click = group["整体点击次数_num"].fillna(0).sum()
        order = group["整体成交订单数_num"].fillna(0).sum()

        created_at = group["素材创建时间"].dropna().min()
        name_values = unique_nonempty(group["素材名称"].tolist())
        title = name_values[0] if name_values else ""
        make_name_values = unique_nonempty(group.get("奥创视频制作名称", pd.Series(dtype=str)).tolist())
        make_name = make_name_values[0] if make_name_values else ""
        maintain_oss_values = unique_nonempty(group.get("维护oss_path", pd.Series(dtype=str)).tolist())
        rows.append(
            {
                "商品": product,
                "账号": account,
                "素材ID": material_id,
                "素材名称": title,
                "素材创建时间": created_at,
                "发布时间": created_at.strftime("%Y-%m-%d") if pd.notna(created_at) else "",
                "奥创视频制作名称": make_name,
                "消耗": spend,
                "GMV": gmv,
                "播放": plays,
                "展示": show,
                "点击": click,
                "订单": order,
                "ROI": (gmv / spend) if spend else None,
                "千次播放GMV": (gmv / plays * 1000) if plays else None,
                "CTR": (click / show) if show else None,
                "CVR": (order / click) if click else None,
                "3s完播率": weighted_rate(group, "3秒播放率_num"),
                "5s完播率": weighted_rate(group, "5秒播放率_num"),
                "10s完播率": weighted_rate(group, "10秒播放率_num"),
                "维护oss_path": maintain_oss_values[0] if maintain_oss_values else "",
            }
        )

    result = pd.DataFrame(rows)
    result = result.sort_values(["商品", "账号", "素材创建时间", "素材ID"]).reset_index(drop=True)
    return result


def classify_result(product: str, gmv: float | None) -> str:
    value = 0 if gmv is None else float(gmv)
    if value == 0:
        return "直接放弃"
    if product == "回味粉":
        return "可复制" if value >= 500 else "需修改"
    return "可复制" if value >= 100 else "需修改"


def build_dashboard(detail: pd.DataFrame) -> pd.DataFrame:
    def make_row(sub: pd.DataFrame, product: str, account: str) -> Dict[str, Any]:
        video_count = int(sub["素材ID"].nunique())
        total_gmv = float(sub["GMV"].fillna(0).sum())
        total_spend = float(sub["消耗"].fillna(0).sum())
        total_play = float(sub["播放"].fillna(0).sum())
        total_show = float(sub["展示"].fillna(0).sum())
        total_click = float(sub["点击"].fillna(0).sum())
        total_order = float(sub["订单"].fillna(0).sum())
        row = {
            "商品": product,
            "账号": account,
            "视频条数": video_count,
            "总GMV": total_gmv,
            "平均GMV": (total_gmv / video_count) if video_count else None,
            "总消耗": total_spend,
            "平均消耗": (total_spend / video_count) if video_count else None,
            "总播放": total_play,
            "平均播放": (total_play / video_count) if video_count else None,
            "总ROI": (total_gmv / total_spend) if total_spend else None,
            "加权CTR": (total_click / total_show) if total_show else None,
            "加权CVR": (total_order / total_click) if total_click else None,
        }
        for col in ["3s完播率", "5s完播率", "10s完播率"]:
            valid = sub[[col, "播放"]].dropna().copy()
            row[f"加权{col}"] = None if valid.empty or valid["播放"].sum() == 0 else float((valid[col] * valid["播放"]).sum() / valid["播放"].sum())
        return row

    rows: List[Dict[str, Any]] = []
    for product in TARGET_PRODUCTS:
        product_df = detail[detail["商品"] == product].copy()
        if product_df.empty:
            continue
        for account in ["奥创", "轩辕"]:
            sub = product_df[product_df["账号"] == account]
            if sub.empty:
                continue
            rows.append(make_row(sub, product, account))
        rows.append(make_row(product_df, product, "总计"))
    if not detail.empty:
        rows.append(make_row(detail, "全部", "总计"))
    return pd.DataFrame(rows)


def load_call_gemini_module() -> Any:
    spec = importlib.util.spec_from_file_location("skill_call_gemini", SKILL_CALL_GEMINI)
    if spec is None or spec.loader is None:
        raise RuntimeError(f"无法加载 Gemini wrapper: {SKILL_CALL_GEMINI}")
    module = importlib.util.module_from_spec(spec)
    spec.loader.exec_module(module)
    return module


class GeminiFallbackClient:
    def run_batch(self, tasks: List[Dict[str, Any]], max_workers: int) -> List[Dict[str, Any]]:
        module = load_call_gemini_module()
        return module.run_video_chat_batch(
            tasks=tasks,
            max_workers=max(1, max_workers),
            retries=2,
            retry_base_delay=1.0,
            submit_qps=1.0,
            poll_qps=4.0,
        )


def asr_schema() -> Dict[str, Any]:
    return {
        "type": "OBJECT",
        "properties": {"asr": {"type": "STRING"}},
        "required": ["asr"],
    }


def asr_prompt(material_id: str, title: str) -> str:
    return (
        f"这是素材ID为 {material_id} 的餐饮千川视频，标题是：{title}。\n"
        "请完整识别视频里的中文口播文字，输出到 asr 字段。\n"
        "要求：尽量保留原顺序、口语停顿和语气词；听不清的少量内容可以结合上下文补齐，但不要编造。"
    )


def hash_key(text: str) -> str:
    return hashlib.md5(text.encode("utf-8")).hexdigest()


def read_json(path: Path) -> Dict[str, Any]:
    return json.loads(path.read_text(encoding="utf-8"))


def write_json(path: Path, payload: Dict[str, Any]) -> None:
    path.write_text(json.dumps(payload, ensure_ascii=False, indent=2), encoding="utf-8")


def normalize_gemini_payload(payload: Any) -> Dict[str, Any]:
    current = payload
    while isinstance(current, dict):
        if "message" in current and isinstance(current["message"], dict):
            current = current["message"]
            continue
        if "result" in current and isinstance(current["result"], dict):
            current = current["result"]
            continue
        if "data" in current and isinstance(current["data"], dict):
            current = current["data"]
            continue
        break
    return current if isinstance(current, dict) else {"_raw": clean_text(current)}


def build_asr_tasks(detail: pd.DataFrame) -> List[Dict[str, Any]]:
    tasks: List[Dict[str, Any]] = []
    for row in detail.itertuples(index=False):
        oss_path = clean_text(getattr(row, "oss_path", ""))
        if not oss_path:
            continue
        tasks.append(
            {
                "素材ID": str(row.素材ID),
                "oss_path": oss_path,
                "素材名称": clean_text(row.素材名称),
                "task": {
                    "proxy_base": PROXY_BASE,
                    "tos_base": TOS_BASE,
                    "model_name": MODEL_NAME,
                    "system_prompt": "你是中文短视频ASR助手，擅长准确识别餐饮视频里的口播内容。",
                    "user_message": asr_prompt(str(row.素材ID), clean_text(row.素材名称)),
                    "oss_links": [oss_path],
                    "response_schema": asr_schema(),
                },
            }
        )
    return tasks


def run_asr(detail: pd.DataFrame, cache_dir: Path, max_workers: int, limit: int | None = None) -> pd.DataFrame:
    cache_dir.mkdir(parents=True, exist_ok=True)
    asr_tasks = build_asr_tasks(detail)
    if limit is not None:
        asr_tasks = asr_tasks[:limit]

    results: Dict[str, Dict[str, str]] = {}
    pending_tasks: List[Dict[str, Any]] = []
    pending_meta: List[Dict[str, Any]] = []

    for item in asr_tasks:
        material_id = item["素材ID"]
        cache_path = cache_dir / f"{material_id}_{hash_key(item['oss_path'])}.json"
        if cache_path.exists():
            cached = read_json(cache_path)
            if not clean_text(cached.get("error")):
                payload = normalize_gemini_payload(cached.get("payload", {}))
                results[material_id] = {"asr": clean_text(payload.get("asr")), "asr_error": ""}
                continue
        pending_tasks.append(item["task"])
        pending_meta.append({"素材ID": material_id, "cache_path": cache_path})

    if pending_tasks:
        try:
            module = load_call_gemini_module()
            batch_results = module.run_video_chat_batch(
                tasks=pending_tasks,
                max_workers=max(1, max_workers),
                retries=3,
                retry_base_delay=1.0,
                submit_qps=1.0,
                poll_qps=4.0,
            )
        except Exception:
            batch_results = GeminiFallbackClient().run_batch(pending_tasks, max_workers=max_workers)

        for meta, result in zip(pending_meta, batch_results):
            payload = result.get("result", {})
            error = clean_text(result.get("error", "")) if not result.get("ok") else ""
            write_json(meta["cache_path"], {"payload": payload, "error": error, "raw_result": result})
            parsed = normalize_gemini_payload(payload)
            results[meta["素材ID"]] = {"asr": clean_text(parsed.get("asr")), "asr_error": error}

    out = detail.copy()
    out["asr"] = out["素材ID"].map(lambda x: results.get(str(x), {}).get("asr", ""))
    out["asr_error"] = out["素材ID"].map(lambda x: results.get(str(x), {}).get("asr_error", ""))
    out.loc[out["oss_path"].isna() | (out["oss_path"].astype(str).str.strip() == ""), "asr_error"] = "缺少oss_path"
    return out


def format_ratio(value: Any) -> Any:
    if value is None or (isinstance(value, float) and pd.isna(value)):
        return ""
    return float(value)


def format_detail(detail: pd.DataFrame) -> pd.DataFrame:
    out = detail.copy()
    out["素材创建时间"] = out["素材创建时间"].dt.strftime("%Y-%m-%d %H:%M:%S")
    out["三类结果"] = out.apply(lambda row: classify_result(str(row["商品"]), row["GMV"]), axis=1)
    out["播放"] = out["播放"].round(0)
    for col in ["消耗", "GMV", "ROI", "千次播放GMV", "CTR", "CVR", "3s完播率", "5s完播率", "10s完播率"]:
        out[col] = out[col].map(format_ratio)

    first_cols = [
        "商品",
        "账号",
        "素材ID",
        "素材名称",
        "奥创视频制作名称",
        "素材创建时间",
        "发布时间",
        "oss_path",
        "asr",
        "asr_error",
        "消耗",
        "GMV",
        "ROI",
        "播放",
        "千次播放GMV",
        "展示",
        "点击",
        "订单",
        "CTR",
        "CVR",
        "3s完播率",
        "5s完播率",
        "10s完播率",
        "三类结果",
    ]
    return out[first_cols].sort_values(["商品", "账号", "三类结果", "GMV", "播放"], ascending=[True, True, True, False, False]).reset_index(drop=True)


def format_dashboard_df(board: pd.DataFrame) -> pd.DataFrame:
    out = board.copy()
    for col in [
        "总GMV",
        "平均GMV",
        "总消耗",
        "平均消耗",
        "总播放",
        "平均播放",
        "总ROI",
        "加权CTR",
        "加权CVR",
        "加权3s完播率",
        "加权5s完播率",
        "加权10s完播率",
    ]:
        out[col] = out[col].map(format_ratio)
    return out


def write_excel(detail: pd.DataFrame, dashboard: pd.DataFrame, output_xlsx: Path) -> None:
    copy_df = detail[detail["三类结果"] == "可复制"].copy()
    modify_df = detail[detail["三类结果"] == "需修改"].copy()
    drop_df = detail[detail["三类结果"] == "直接放弃"].copy()

    with pd.ExcelWriter(output_xlsx, engine="openpyxl") as writer:
        detail.to_excel(writer, index=False, sheet_name="总明细表")
        dashboard.to_excel(writer, index=False, sheet_name="看板数据表")
        copy_df.to_excel(writer, index=False, sheet_name="可复制")
        modify_df.to_excel(writer, index=False, sheet_name="需修改")
        drop_df.to_excel(writer, index=False, sheet_name="直接放弃")

    wb = load_workbook(output_xlsx)
    wrap = Alignment(wrap_text=True, vertical="top")
    bold = Font(bold=True)
    fill = PatternFill("solid", fgColor="D9EAF7")
    width_map = {
        "总明细表": {
            "A": 10, "B": 8, "C": 18, "D": 52, "E": 18, "F": 20, "G": 12, "H": 68, "I": 72, "J": 20,
            "K": 12, "L": 12, "M": 12, "N": 12, "O": 14, "P": 12, "Q": 10, "R": 10, "S": 12, "T": 12,
            "U": 12, "V": 12, "W": 12, "X": 12,
        },
        "看板数据表": {
            "A": 10, "B": 10, "C": 10, "D": 12, "E": 12, "F": 12, "G": 12, "H": 12, "I": 12, "J": 12,
            "K": 12, "L": 14, "M": 14, "N": 14,
        },
        "可复制": {"D": 52, "H": 68, "I": 72},
        "需修改": {"D": 52, "H": 68, "I": 72},
        "直接放弃": {"D": 52, "H": 68, "I": 72},
    }

    for ws in wb.worksheets:
        for cell in ws[1]:
            cell.font = bold
            cell.fill = fill
            cell.alignment = wrap
        for row in ws.iter_rows(min_row=2):
            for cell in row:
                cell.alignment = wrap
        for col, width in width_map.get(ws.title, {}).items():
            ws.column_dimensions[col].width = width

    wb.save(output_xlsx)


def write_summary(
    output_md: Path,
    b_date: pd.Timestamp,
    publish_start: pd.Timestamp,
    publish_end: pd.Timestamp,
    detail: pd.DataFrame,
    duplicate_oss_count: int,
) -> None:
    lines = [
        "# 上上周新创建素材分析（workflow版）",
        "",
        f"- b日期：`{b_date.strftime('%Y-%m-%d')}`",
        f"- 上上周发布日期窗口：`{publish_start.strftime('%Y-%m-%d')} ~ {publish_end.strftime('%Y-%m-%d')}`",
        f"- 上上周素材数：`{detail['素材ID'].nunique()}`",
        f"- OSS映射重复视频ID数：`{duplicate_oss_count}`",
        f"- 有oss_path素材数：`{int(detail['oss_path'].fillna('').astype(str).str.strip().ne('').sum())}`",
        f"- ASR成功数：`{int(detail['asr'].fillna('').astype(str).str.strip().ne('').sum())}`",
        f"- ASR失败/缺失数：`{int(detail['asr_error'].fillna('').astype(str).str.strip().ne('').sum())}`",
        "",
        "## 三类结果数量",
    ]
    counts = detail["三类结果"].value_counts().to_dict()
    for label in ["可复制", "需修改", "直接放弃"]:
        lines.append(f"- {label}：`{counts.get(label, 0)}`")
    output_md.write_text("\n".join(lines), encoding="utf-8")


def parse_args() -> argparse.Namespace:
    parser = argparse.ArgumentParser(description="按 workflow 生成上上周新创建素材分析结果")
    parser.add_argument("--max-workers", type=int, default=2, help="Gemini ASR 并发数")
    parser.add_argument("--asr-limit", type=int, default=None, help="仅测试前N条ASR")
    parser.add_argument("--skip-asr", action="store_true", help="跳过ASR")
    return parser.parse_args()


def main() -> None:
    args = parse_args()
    OUTPUT_DIR.mkdir(parents=True, exist_ok=True)

    raw = pd.concat(
        [
            load_recent30(AOCHUANG_XLSX, "奥创"),
            load_recent30(YUANYUAN_XLSX, "轩辕"),
        ],
        ignore_index=True,
    )
    maintain = load_maintenance()
    raw = apply_maintenance(raw, maintain)

    daily_dates = raw.loc[(raw["日期_raw"] != "全部") & raw["日期"].notna(), "日期"]
    if daily_dates.empty:
        raise RuntimeError("最近30天数据里没有有效的消耗日期")
    b_date = daily_dates.max().normalize()
    publish_start = b_date - pd.Timedelta(days=14)
    publish_end = b_date - pd.Timedelta(days=7)

    detail = aggregate_videos(raw, publish_start, publish_end)
    if detail.empty:
        raise RuntimeError("上上周窗口内没有筛到素材")

    oss_map, duplicate_oss_count = load_oss_map()
    detail = detail.merge(oss_map, on="素材ID", how="left")
    detail["oss_path"] = detail["oss_path"].where(detail["oss_path"].notna(), detail["维护oss_path"])
    detail = detail.drop(columns=["维护oss_path"])

    if args.skip_asr:
        detail["asr"] = ""
        detail["asr_error"] = detail["oss_path"].fillna("").map(lambda x: "" if clean_text(x) else "缺少oss_path")
    else:
        cache_dir = OUTPUT_DIR / "workflow_asr_cache_0405-0412"
        detail = run_asr(detail, cache_dir=cache_dir, max_workers=args.max_workers, limit=args.asr_limit)

    formatted_detail = format_detail(detail)
    dashboard = format_dashboard_df(build_dashboard(detail))

    date_label = f"{publish_start.strftime('%m%d')}-{publish_end.strftime('%m%d')}_消耗{(b_date - pd.Timedelta(days=29)).strftime('%m%d')}-{b_date.strftime('%m%d')}"
    output_xlsx = OUTPUT_DIR / f"上上周新创建素材_workflow结果_{date_label}.xlsx"
    output_csv = OUTPUT_DIR / f"上上周新创建素材_workflow结果_{date_label}.csv"
    output_md = OUTPUT_DIR / f"上上周新创建素材_workflow结果_{date_label}.md"

    write_excel(formatted_detail, dashboard, output_xlsx)
    formatted_detail.to_csv(output_csv, index=False, encoding="utf-8-sig", quoting=csv.QUOTE_MINIMAL)
    write_summary(output_md, b_date, publish_start, publish_end, formatted_detail, duplicate_oss_count)

    print(output_xlsx)
    print(output_csv)
    print(output_md)


if __name__ == "__main__":
    main()
